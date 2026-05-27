import pandas as pd
import openpyxl
import logging

class XLSXWriter:
    """Write DataFrames to XLSX following SGAcloud templates"""
    
    def write_products(self, df: pd.DataFrame, template_path: str, output_path: str):
        """Write products DataFrame to MERCADORIAS template - data starts at row 3, column 2"""
        try:
            # Load template
            template_wb = openpyxl.load_workbook(template_path)
            ws = template_wb[template_wb.sheetnames[0]]
            
            # Clear existing data rows starting from row 3 (keep rows 1-2)
            for row in range(3, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).value = None
            
            # Write data starting from row 3, column 2 (col 1 = COLUNA stays empty)
            for idx, row in enumerate(df.itertuples(index=False), start=3):
                # Write each column: DataFrame col i -> Template col (i+2)
                for col_idx, value in enumerate(row, start=2):
                    ws.cell(row=idx, column=col_idx, value=value)
            
            # Save
            template_wb.save(output_path)
            logging.info(f'Written {len(df)} product records to {output_path}')
        except Exception as e:
            logging.error(f'Error writing XLSX: {e}')
            raise
    
    def write_clients(self, df: pd.DataFrame, template_path: str, output_path: str):
        """Write clients/suppliers DataFrame to CLIENTES template - data starts at row 4, column 2"""
        try:
            # Load template
            template_wb = openpyxl.load_workbook(template_path)
            ws = template_wb[template_wb.sheetnames[0]]
            
            # Clear existing data rows starting from row 4 (keep rows 1-3)
            for row in range(4, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).value = None
            
            # Write data starting from row 4, column 2 (col 1 = COLUNA stays empty)
            for idx, row in enumerate(df.itertuples(index=False), start=4):
                # Write each column: DataFrame col i -> Template col (i+2)
                for col_idx in range(len(row)):
                    value = row[col_idx]
                    # Template column = col_idx + 2 (skip col 1 which is COLUNA)
                    ws.cell(row=idx, column=col_idx+2, value=value)
            
            # Save
            template_wb.save(output_path)
            logging.info(f'Written {len(df)} client/supplier records to {output_path}')
        except Exception as e:
            logging.error(f'Error writing XLSX: {e}')
            raise
