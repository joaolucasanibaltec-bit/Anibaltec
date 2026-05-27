import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import xlrd
from openpyxl import load_workbook
import shutil
from datetime import datetime
import os
import sys
import csv

# Banco de dados IBGE
import unicodedata

def remove_accents(text):
    if not text:
        return ''
    text = text.upper().strip()
    nfkd = unicodedata.normalize('NFKD', text)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ARQUIVO_IBGE = os.path.join(BASE_DIR, 'ibge_municipios.csv')

_cache_ibge = None

def carregar_ibge():
    global _cache_ibge
    if _cache_ibge is not None:
        return _cache_ibge
    
    _cache_ibge = {}
    try:
        with open(ARQUIVO_IBGE, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader)
            for row in reader:
                if len(row) >= 4:
                    codigo = row[0]
                    uf_codigo = row[1]
                    nome = row[3]
                    
                    uf_map = {'11': 'RO', '12': 'AC', '13': 'AM', '14': 'RR', '15': 'PA', '16': 'AP', '17': 'TO',
                              '21': 'MA', '22': 'PI', '23': 'CE', '24': 'RN', '25': 'PB', '26': 'PE', '27': 'AL', '28': 'SE',
                              '29': 'BA', '31': 'MG', '32': 'ES', '33': 'RJ', '35': 'SP',
                              '41': 'PR', '42': 'SC', '43': 'RS',
                              '50': 'MS', '51': 'MT', '52': 'GO', '53': 'DF'}
                    
                    uf = uf_map.get(uf_codigo, uf_codigo)
                    
                    nome_sem_acento = remove_accents(nome)
                    _cache_ibge[(nome_sem_acento, uf)] = codigo
    except Exception as e:
        print(f"Erro ao carregar IBGE: {e}")
    
    return _cache_ibge

def get_codigo_ibge(municipio, uf):
    if not municipio or not uf:
        return ''
    
    dados = carregar_ibge()
    municipio_sem_acento = remove_accents(municipio)
    uf_upper = uf.upper().strip()
    
    key = (municipio_sem_acento, uf_upper)
    if key in dados:
        return dados[key]
    
    for (nome, estado), codigo in dados.items():
        if municipio_sem_acento in nome or nome in municipio_sem_acento:
            if uf_upper == estado:
                return codigo
    return ''

class ConversorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Conversor ACSN para SgaCloud - Anibaltec")
        self.root.geometry("800x650")
        self.root.resizable(False, False)
        
        # Cores Anibaltec
        self.cor_principal = "#0D2C54"  # Azul escuro
        self.cor_secundaria = "#F2951B"  # Laranja
        self.cor_fundo = "#F5F5F5"       # Cinza claro
        self.cor_texto = "#333333"       # Cinza escuro
        self.cor_branco = "#FFFFFF"
        
        self.arquivo_acsn = None
        self.df_acsn = None
        self.modo = "produtos"
        
        # Templates padrão
        if getattr(sys, 'frozen', False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
        
        self.template_mercadorias = os.path.join(base_dir, "Templates", "SgaCloud", "TEMPLATE MERCADORIAS.xlsx")
        self.template_clientes = os.path.join(base_dir, "Templates", "SgaCloud", "TEMPLATE CLIENTES.xlsx")
        
        # Se não encontrou, tenta caminho original (para desenvolvimento)
        if not os.path.exists(self.template_mercadorias):
            self.template_mercadorias = os.path.join(base_dir, "SgaCloud", "TEMPLATE MERCADORIAS.xlsx")
        if not os.path.exists(self.template_clientes):
            self.template_clientes = os.path.join(base_dir, "SgaCloud", "TEMPLATE CLIENTES.xlsx")
        
        self.criar_interface()
    
    def criar_interface(self):
        # Frame do título com gradiente
        frame_topo = tk.Frame(self.root, bg=self.cor_principal, height=90)
        frame_topo.pack(fill='x')
        frame_topo.pack_propagate(False)
        
        # Logo/Título
        lbl_logo = tk.Label(frame_topo, text="ANIBALTEC", 
                          font=("Arial", 22, "bold"), 
                          fg=self.cor_branco, bg=self.cor_principal)
        lbl_logo.pack(pady=(12, 3))
        
        lbl_subtitle = tk.Label(frame_topo, text="Conversor ACSN → SgaCloud",
                               font=("Arial", 11), 
                               fg=self.cor_secundaria, bg=self.cor_principal)
        lbl_subtitle.pack(pady=(0, 8))
        
        # Frame principal com fundo
        frame_main = tk.Frame(self.root, bg=self.cor_fundo)
        frame_main.pack(fill='both', expand=True, padx=20, pady=15)
        
        # Notebook
        self.notebook = ttk.Notebook(frame_main)
        self.notebook.pack(fill='both', expand=True)
        
        # Aba Produtos - com espaços para igualar largura (21 chars)
        self.aba_produtos = tk.Frame(self.notebook, bg=self.cor_fundo, height=400)
        self.notebook.add(self.aba_produtos, text="Produtos             ", padding=15)
        self.criar_aba_produtos()
        
        # Aba Clientes/Fornecedores - 21 chars
        self.aba_clientes = tk.Frame(self.notebook, bg=self.cor_fundo, height=400)
        self.notebook.add(self.aba_clientes, text="Clientes/Fornecedores", padding=15)
        self.criar_aba_clientes()
        
        # Rodapé
        frame_rodape = tk.Frame(self.root, bg=self.cor_principal, height=30)
        frame_rodape.pack(fill='x')
        frame_rodape.pack_propagate(False)
        
        lbl_rodape = tk.Label(frame_rodape, 
                            text="Anibaltec Automação | (84) 3323-8470 | anibaltec.com.br",
                            font=("Arial", 8), 
                            fg="#CCCCCC", bg=self.cor_principal)
        lbl_rodape.pack(pady=6)
    
    def criar_aba_produtos(self):
        frame = self.aba_produtos
        
        # Card principal - preencher todo o espaço
        card = tk.Frame(frame, bg=self.cor_branco, relief="flat", bd=1)
        card.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Título da seção
        titulo_secao = tk.Label(card, text="Converter Produtos ACSN → SgaCloud",
                              font=("Arial", 14, "bold"),
                              fg=self.cor_principal, bg=self.cor_branco)
        titulo_secao.pack(pady=(15, 10))
        
        descricao = tk.Label(card, text="Converta planilhas de produtos para o formato SgaCloud",
                            font=("Arial", 10), fg=self.cor_texto, bg=self.cor_branco)
        descricao.pack(pady=(0, 15))
        
        # Área de seleção
        frame_selecao = tk.Frame(card, bg=self.cor_branco)
        frame_selecao.pack(pady=10)
        
        self.label_arquivo_prod = tk.Label(frame_selecao, 
                                         text="Nenhum arquivo selecionado",
                                         font=("Arial", 10), 
                                         fg="gray", bg=self.cor_branco,
                                         width=35, anchor='w')
        self.label_arquivo_prod.pack(side=tk.LEFT, padx=(0, 10))
        
        btn_selecionar = tk.Button(frame_selecao, text="Selecionar Arquivo",
                                  font=("Arial", 10, "bold"),
                                  command=self.selecionar_arquivo_produtos,
                                  bg=self.cor_principal, fg=self.cor_branco,
                                  padx=20, pady=8, relief="flat", cursor="hand2")
        btn_selecionar.pack(side=tk.LEFT)
        
        # Template fixo
        frame_template = tk.Frame(card, bg=self.cor_branco)
        frame_template.pack(pady=10)
        
        lbl_template_tit = tk.Label(frame_template, text="Template SgaCloud:",
                                   font=("Arial", 9), fg="gray", bg=self.cor_branco)
        lbl_template_tit.pack()
        
        self.label_template_prod = tk.Label(frame_template, 
                                           text=os.path.basename(self.template_mercadorias),
                                           font=("Arial", 10, "bold"), 
                                           fg=self.cor_principal, bg=self.cor_branco)
        self.label_template_prod.pack()
        
        # Info
        self.label_info_prod = tk.Label(card, text="", 
                                       font=("Arial", 10), 
                                       fg=self.cor_texto, bg=self.cor_branco)
        self.label_info_prod.pack(pady=10)
        
        # Botão Converter
        self.btn_converter_prod = tk.Button(card, text="Converter para SgaCloud",
                                      font=("Arial", 12, "bold"),
                                      command=self.converter_produtos,
                                      bg=self.cor_secundaria, fg=self.cor_principal,
                                      padx=40, pady=12, relief="flat", cursor="hand2",
                                      state=tk.DISABLED)
        self.btn_converter_prod.pack(pady=15)
        
        # Barra de progresso
        style = ttk.Style()
        style.configure("Horizontal.TProgressbar", 
                       troughcolor=self.cor_fundo,
                       background=self.cor_secundaria,
                       thickness=20)
        self.progress_prod = ttk.Progressbar(card, length=500, 
                                            mode='determinate',
                                            style="Horizontal.TProgressbar")
        self.progress_prod.pack(pady=10)
        
        # Label resultado
        self.label_resultado_prod = tk.Label(card, text="", 
                                           font=("Arial", 10, "bold"),
                                           fg=self.cor_principal, bg=self.cor_branco)
        self.label_resultado_prod.pack(pady=(0, 15))
    
    def criar_aba_clientes(self):
        frame = self.aba_clientes
        
        # Card principal - preencher todo o espaço
        card = tk.Frame(frame, bg=self.cor_branco, relief="flat", bd=1)
        card.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Título da seção
        titulo_secao = tk.Label(card, text="Converter Clientes e Fornecedores",
                              font=("Arial", 14, "bold"),
                              fg=self.cor_principal, bg=self.cor_branco)
        titulo_secao.pack(pady=(15, 10))
        
        descricao = tk.Label(card, text="Unifique clientes e fornecedores em uma única planilha SgaCloud",
                            font=("Arial", 10), fg=self.cor_texto, bg=self.cor_branco)
        descricao.pack(pady=(0, 15))
        
        # Botões para selecionar arquivos
        frame_botoes = tk.Frame(card, bg=self.cor_branco)
        frame_botoes.pack(pady=10)
        
        btn_cliente = tk.Button(frame_botoes, text="Selecionar Clientes",
                               font=("Arial", 10, "bold"),
                               command=self.selecionar_clientes,
                               bg=self.cor_principal, fg=self.cor_branco,
                               padx=20, pady=8, relief="flat", cursor="hand2")
        btn_cliente.pack(side=tk.LEFT, padx=5)
        
        btn_fornecedor = tk.Button(frame_botoes, text="Selecionar Fornecedores",
                                  font=("Arial", 10, "bold"),
                                  command=self.selecionar_fornecedores,
                                  bg="#7B1FA2", fg=self.cor_branco,
                                  padx=20, pady=8, relief="flat", cursor="hand2")
        btn_fornecedor.pack(side=tk.LEFT, padx=5)
        
        # Labels info
        frame_labels = tk.Frame(card, bg=self.cor_branco)
        frame_labels.pack(pady=10)
        
        self.label_cliente = tk.Label(frame_labels, 
                                      text="Nenhum cliente selecionado", 
                                      font=("Arial", 10), fg="gray", bg=self.cor_branco)
        self.label_cliente.pack(pady=3)
        
        self.label_fornecedor = tk.Label(frame_labels, 
                                         text="Nenhum fornecedor selecionado", 
                                         font=("Arial", 10), fg="gray", bg=self.cor_branco)
        self.label_fornecedor.pack(pady=3)
        
        # Template fixo
        frame_template = tk.Frame(card, bg=self.cor_branco)
        frame_template.pack(pady=10)
        
        lbl_template_tit = tk.Label(frame_template, text="Template SgaCloud:",
                                   font=("Arial", 9), fg="gray", bg=self.cor_branco)
        lbl_template_tit.pack()
        
        self.label_template_cli = tk.Label(frame_template, 
                                           text=os.path.basename(self.template_clientes),
                                           font=("Arial", 10, "bold"), 
                                           fg=self.cor_principal, bg=self.cor_branco)
        self.label_template_cli.pack()
        
        # Botão Converter
        self.btn_conv_cli = tk.Button(card, text="Converter para SgaCloud",
                                      font=("Arial", 12, "bold"),
                                      command=self.converter_clientes,
                                      bg=self.cor_secundaria, fg=self.cor_principal,
                                      padx=40, pady=12, relief="flat", cursor="hand2",
                                      state=tk.DISABLED)
        self.btn_conv_cli.pack(pady=15)
        
        # Barra de progresso
        style = ttk.Style()
        style.configure("Horizontal.TProgressbar", 
                       troughcolor=self.cor_fundo,
                       background=self.cor_secundaria,
                       thickness=20)
        self.progress_cli = ttk.Progressbar(card, length=500, 
                                            mode='determinate',
                                            style="Horizontal.TProgressbar")
        self.progress_cli.pack(pady=10)
        
        # Label resultado
        self.label_resultado_cli = tk.Label(card, text="", 
                                           font=("Arial", 10, "bold"),
                                           fg=self.cor_principal, bg=self.cor_branco)
        self.label_resultado_cli.pack(pady=(0, 15))
    
    # === PRODUTOS ===
    def selecionar_arquivo_produtos(self):
        arquivo = filedialog.askopenfilename(
            title="Selecionar Planilha ACSN - PRODUTOS",
            filetypes=[("Arquivos XLS", "*.XLS"), ("Todos os arquivos", "*.*")]
        )
        
        if arquivo:
            self.arquivo_acsn = arquivo
            nome = os.path.basename(arquivo)
            self.label_arquivo_prod.config(text=nome, fg="black")
            self.btn_converter_prod.config(state=tk.NORMAL)
            self.label_resultado_prod.config(text="")
            
            try:
                book = xlrd.open_workbook(arquivo, encoding_override='latin1')
                sheet = book.sheet_by_index(0)
                total = sheet.nrows - 1
                self.label_info_prod.config(text=f"Planilha carregada: {total} produtos encontrados")
            except Exception as e:
                self.label_info_prod.config(text=f"Erro ao ler arquivo: {str(e)}", fg="red")
    
    def converter_produtos(self):
        if not self.arquivo_acsn:
            messagebox.showerror("Erro", "Selecione uma planilha primeiro!")
            return
        
        template_path = self.template_mercadorias
        
        if not os.path.exists(template_path):
            messagebox.showerror("Erro", f"Template não encontrado: {template_path}")
            return
        
        try:
            self.btn_converter_prod.config(state=tk.DISABLED)
            self.progress_prod['value'] = 10
            self.root.update()
            
            # Ler dados ACSN
            book = xlrd.open_workbook(self.arquivo_acsn, encoding_override='latin1')
            sheet = book.sheet_by_index(0)
            headers = [sheet.cell_value(0, j) for j in range(sheet.ncols)]
            dados = []
            
            for i in range(1, sheet.nrows):
                row = {}
                for j, header in enumerate(headers):
                    cell = sheet.cell(i, j)
                    if cell.ctype == xlrd.XL_CELL_EMPTY:
                        row[header] = None
                    elif cell.ctype == xlrd.XL_CELL_TEXT:
                        row[header] = sheet.cell_value(i, j).strip()
                    else:
                        row[header] = sheet.cell_value(i, j)
                dados.append(row)
                if i % 500 == 0:
                    self.progress_prod['value'] = 10 + (i / sheet.nrows) * 30
                    self.root.update()
            
            self.df_acsn = pd.DataFrame(dados)
            self.progress_prod['value'] = 40
            self.root.update()
            
            # Copiar template
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            dir_output = os.path.dirname(template_path)
            saida = os.path.join(dir_output, 'output', f'MERCADORIAS_SGACLOUD_{ts}.xlsx')
            os.makedirs(os.path.dirname(saida), exist_ok=True)
            
            shutil.copy(template_path, saida)
            self.progress_prod['value'] = 50
            self.root.update()
            
            # Abrir workbook
            wb = load_workbook(saida)
            ws = wb.active
            
            # Mapeamento produtos
            mapeamento = {
                2: 'procod',
                3: 'prodes',
                5: 'prouni',
                6: 'clf_ncm',
                7: 'procest',
                8: 'propdv',
                9: 'procsosn',
                12: 'propdc',
                14: 'proref',
                24: 'propis',
                25: 'prostipi',
                26: 'procodipi',
                28: 'proorig',
                29: 'proipi',
                35: 'proqmi',
            }
            
            conversao_csosn = {'102': '00', '500': '60', '900': '90'}
            
            def formatar_valor(valor, campo):
                if valor is None or (isinstance(valor, float) and valor != valor):
                    return ''
                if isinstance(valor, float):
                    if campo in ['procod', 'clf_ncm', 'procest', 'proorig']:
                        return str(int(valor))
                    if valor == int(valor):
                        return int(valor)
                    return round(valor, 4)
                if campo == 'procsosn':
                    return conversao_csosn.get(str(valor), str(valor))
                return str(valor).strip()
            
            linha_inicio = 3
            total = len(self.df_acsn)
            
            for idx, row in self.df_acsn.iterrows():
                for col_num, campo_acsn in mapeamento.items():
                    valor = row.get(campo_acsn)
                    valor_formatado = formatar_valor(valor, campo_acsn)
                    ws.cell(row=linha_inicio + idx, column=col_num, value=valor_formatado)
                
                if idx % 500 == 0:
                    self.progress_prod['value'] = 50 + (idx / total) * 45
                    self.root.update()
            
            wb.save(saida)
            self.progress_prod['value'] = 100
            
            self.label_resultado_prod.config(text=f"Sucesso! {total} produtos convertidos!", fg="green")
            os.startfile(saida)
            
            messagebox.showinfo("Sucesso", f"Planilha gerada!\n{os.path.basename(saida)}\n{total} produtos.")
            self.btn_converter_prod.config(state=tk.NORMAL)
            
        except Exception as e:
            self.btn_converter_prod.config(state=tk.NORMAL)
            self.progress_prod['value'] = 0
            messagebox.showerror("Erro", f"Erro na conversão:\n{str(e)}")
    
    # === CLIENTES ===
    def selecionar_clientes(self):
        arquivo = filedialog.askopenfilename(
            title="Selecionar Clientes ACSN",
            filetypes=[("Arquivos XLS", "*.XLS"), ("Todos os arquivos", "*.*")]
        )
        
        if arquivo:
            self.arquivo_cliente = arquivo
            self.label_cliente.config(text=os.path.basename(arquivo), fg="black")
            self.verificar_botoes_clientes()
    
    def selecionar_fornecedores(self):
        arquivo = filedialog.askopenfilename(
            title="Selecionar Fornecedores ACSN",
            filetypes=[("Arquivos XLS", "*.XLS"), ("Todos os arquivos", "*.*")]
        )
        
        if arquivo:
            self.arquivo_fornecedor = arquivo
            self.label_fornecedor.config(text=os.path.basename(arquivo), fg="black")
            self.verificar_botoes_clientes()
    
    def verificar_botoes_clientes(self):
        if hasattr(self, 'arquivo_cliente') and hasattr(self, 'arquivo_fornecedor'):
            self.btn_conv_cli.config(state=tk.NORMAL)
    
    def converter_clientes(self):
        template_path = self.template_clientes
        
        if not os.path.exists(template_path):
            messagebox.showerror("Erro", f"Template não encontrado: {template_path}")
            return
        
        # Criar copia temporaria do template para evitar erro de permissao
        import tempfile
        dir_temp = tempfile.gettempdir()
        nome_template = os.path.basename(template_path)
        template_copia = os.path.join(dir_temp, f'copia_template_{nome_template}')
        
        try:
            if os.path.exists(template_copia):
                os.remove(template_copia)
        except:
            pass
        
        shutil.copy(template_path, template_copia)
        template_path = template_copia
        
        try:
            self.btn_conv_cli.config(state=tk.DISABLED)
            self.progress_cli['value'] = 10
            self.root.update()
            
            # Ler CLIENTES
            dados_total = []
            
            if hasattr(self, 'arquivo_cliente'):
                book = xlrd.open_workbook(self.arquivo_cliente, encoding_override='latin1')
                sheet = book.sheet_by_index(0)
                headers = [sheet.cell_value(0, j) for j in range(sheet.ncols)]
                
                for i in range(1, sheet.nrows):
                    row = {}
                    for j, header in enumerate(headers):
                        cell = sheet.cell(i, j)
                        if cell.ctype == xlrd.XL_CELL_EMPTY:
                            row[header] = None
                        elif cell.ctype == xlrd.XL_CELL_TEXT:
                            row[header] = sheet.cell_value(i, j).strip()
                        else:
                            row[header] = sheet.cell_value(i, j)
                    row['_tipo'] = 'C'  # Cliente
                    dados_total.append(row)
            
            self.progress_cli['value'] = 30
            self.root.update()
            
            # Ler FORNECEDORES
            if hasattr(self, 'arquivo_fornecedor'):
                book = xlrd.open_workbook(self.arquivo_fornecedor, encoding_override='latin1')
                sheet = book.sheet_by_index(0)
                headers = [sheet.cell_value(0, j) for j in range(sheet.ncols)]
                
                for i in range(1, sheet.nrows):
                    row = {}
                    for j, header in enumerate(headers):
                        cell = sheet.cell(i, j)
                        if cell.ctype == xlrd.XL_CELL_EMPTY:
                            row[header] = None
                        elif cell.ctype == xlrd.XL_CELL_TEXT:
                            row[header] = sheet.cell_value(i, j).strip()
                        else:
                            row[header] = sheet.cell_value(i, j)
                    row['_tipo'] = 'F'  # Fornecedor
                    dados_total.append(row)
            
            self.progress_cli['value'] = 50
            self.root.update()
            
            df_dados = pd.DataFrame(dados_total)
            total = len(df_dados)
            
            # Copiar template
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            dir_output = os.path.dirname(template_path)
            saida = os.path.join(dir_output, 'output', f'CLIENTE_FORNECEDOR_SGACLOUD_{ts}.xlsx')
            os.makedirs(os.path.dirname(saida), exist_ok=True)
            
            shutil.copy(template_path, saida)
            
            # Abrir workbook
            wb = load_workbook(saida)
            ws = wb.active
            
            # Conversor de data
            meses = {'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04', 'may': '05', 'jun': '06',
                     'jul': '07', 'aug': '08', 'sep': '09', 'oct': '10', 'nov': '11', 'dec': '12'}
            
            def converter_data(data_str):
                if not data_str or data_str == '':
                    return ''
                data_str = str(data_str).strip().lower()
                try:
                    parts = data_str.replace('-', ' ').split()
                    if len(parts) == 3:
                        dia, mes, ano = parts
                        if mes in meses:
                            ano_full = '20' + ano if len(ano) == 2 else ano
                            return f'{dia}/{meses[mes]}/{ano_full}'
                except:
                    pass
                return ''
            
            def get_ibge(municipio, uf):
                return get_codigo_ibge(municipio, uf)
            # Coluna: Campo ACSN
            mapeamento = {
                2: 'fccod',      # codigo
                3: 'fcnom',      # razao
                4: 'fcnom',      # fantasia
                5: 'fccpg',      # cpf cnpj
                6: '_tipo',      # tipo (C ou F)
                7: 'fcrg',      # rg ie
                9: 'fcend',     # logradouro
                10: 'SN',       # numero
                11: 'fccom',     # complemento
                12: 'fccep',     # cep
                13: 'fcbai',     # bairro
                15: 'fccid',     # municipio
                16: 'fcest',     # uf
                17: 'fccad',     # cadastro
                19: 'fcemail',   # email
                20: 'fcfon',    # fone resid
                23: 'fclimcr',   # limite
            }
            
            def formatar_valor_cli(valor):
                if valor is None or (isinstance(valor, float) and valor != valor):
                    return ''
                if isinstance(valor, float):
                    if valor == int(valor):
                        return int(valor)
                    return round(valor, 2)
                return str(valor).strip()
            
            # Inserir dados a partir da linha 4 (linha 3 = dados)
            linha_inicio = 4
            
            for idx, row in df_dados.iterrows():
                # Coluna 2: codigo
                ws.cell(row=linha_inicio + idx, column=2, value=formatar_valor_cli(row.get('fccod')))
                # Coluna 3: razao
                ws.cell(row=linha_inicio + idx, column=3, value=formatar_valor_cli(row.get('fcnom')).upper())
                # Coluna 4: fantasia
                ws.cell(row=linha_inicio + idx, column=4, value=formatar_valor_cli(row.get('fcnom')).upper())
                # Coluna 5: cpf cnpj
                ws.cell(row=linha_inicio + idx, column=5, value=formatar_valor_cli(row.get('fccpg')))
                # Coluna 6: tipo (C ou F)
                ws.cell(row=linha_inicio + idx, column=6, value=row.get('_tipo', ''))
                # Coluna 7: rg ie
                ws.cell(row=linha_inicio + idx, column=7, value=formatar_valor_cli(row.get('fcrg')))
                # Coluna 9: logradouro
                ws.cell(row=linha_inicio + idx, column=9, value=formatar_valor_cli(row.get('fcend')))
                # Coluna 10: numero
                ws.cell(row=linha_inicio + idx, column=10, value='SN')
                # Coluna 11: complemento
                ws.cell(row=linha_inicio + idx, column=11, value=formatar_valor_cli(row.get('fccom')))
                # Coluna 12: cep
                ws.cell(row=linha_inicio + idx, column=12, value=formatar_valor_cli(row.get('fccep')))
                # Coluna 13: bairro
                ws.cell(row=linha_inicio + idx, column=13, value=formatar_valor_cli(row.get('fcbai')))
                # Coluna 15: municipio
                ws.cell(row=linha_inicio + idx, column=15, value=formatar_valor_cli(row.get('fccid')).upper())
                # Coluna 14: ibge
                ws.cell(row=linha_inicio + idx, column=14, value=get_ibge(row.get('fccid', ''), row.get('fcest', '')))
                # Coluna 16: uf
                ws.cell(row=linha_inicio + idx, column=16, value=formatar_valor_cli(row.get('fcest')))
                # Coluna 17: cadastro (formatar data)
                data_cadastro = converter_data(row.get('fccad', ''))
                ws.cell(row=linha_inicio + idx, column=17, value=data_cadastro if data_cadastro else formatar_valor_cli(row.get('fccad')))
                # Coluna 19: email
                ws.cell(row=linha_inicio + idx, column=19, value=formatar_valor_cli(row.get('fcemail')))
                # Coluna 20: fone resid
                ws.cell(row=linha_inicio + idx, column=20, value=formatar_valor_cli(row.get('fcfon')))
                # Coluna 23: limite
                ws.cell(row=linha_inicio + idx, column=23, value=formatar_valor_cli(row.get('fclimcr')))
                # Coluna 25: dt vencimento (padrão)
                ws.cell(row=linha_inicio + idx, column=25, value='01/01/1900')
                
                if idx % 100 == 0:
                    self.progress_cli['value'] = 50 + (idx / total) * 45
                    self.root.update()
            
            wb.save(saida)
            self.progress_cli['value'] = 100
            
            clientes = len([r for r in dados_total if r.get('_tipo') == 'C'])
            fornecedores = len([r for r in dados_total if r.get('_tipo') == 'F'])
            
            self.label_resultado_cli.config(
                text=f"Sucesso! {clientes} clientes + {fornecedores} fornecedores",
                fg="green"
            )
            os.startfile(saida)
            
            messagebox.showinfo("Sucesso", 
                f"Planilha gerada!\n{os.path.basename(saida)}\n{clientes} clientes\n{fornecedores} fornecedores.")
            self.btn_conv_cli.config(state=tk.NORMAL)
            
        except Exception as e:
            self.btn_conv_cli.config(state=tk.NORMAL)
            self.progress_cli['value'] = 0
            messagebox.showerror("Erro", f"Erro na conversão:\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ConversorApp(root)
    root.mainloop()