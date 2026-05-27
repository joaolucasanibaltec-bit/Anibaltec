import os
import shutil
import csv
import unicodedata
from pathlib import Path
from datetime import datetime
import xlrd
import pandas as pd
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent.parent

def _remove_accents(text):
    if not text:
        return ""
    text = text.upper().strip()
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))

class IBGELookup:
    _cache = None

    def get(self, municipio, uf):
        if not municipio or not uf:
            return ""
        if self._cache is None:
            self._load()
        key = (_remove_accents(municipio), uf.upper().strip())
        if key in self._cache:
            return self._cache[key]
        for (nome, estado), codigo in self._cache.items():
            if key[0] in nome or nome in key[0]:
                if key[1] == estado:
                    return codigo
        return ""

    def _load(self):
        self._cache = {}
        ibge_path = BASE / "ibge_municipios.csv"
        if not ibge_path.exists():
            return
        uf_map = {
            "11": "RO", "12": "AC", "13": "AM", "14": "RR", "15": "PA", "16": "AP", "17": "TO",
            "21": "MA", "22": "PI", "23": "CE", "24": "RN", "25": "PB", "26": "PE", "27": "AL", "28": "SE",
            "29": "BA", "31": "MG", "32": "ES", "33": "RJ", "35": "SP",
            "41": "PR", "42": "SC", "43": "RS", "50": "MS", "51": "MT", "52": "GO", "53": "DF",
        }
        try:
            with open(ibge_path, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                next(reader)
                for row in reader:
                    if len(row) >= 4:
                        nome = _remove_accents(row[3])
                        uf = uf_map.get(row[1], row[1])
                        self._cache[(nome, uf)] = row[0]
        except Exception:
            pass

class ConverterService:
    def __init__(self):
        self.ibge = IBGELookup()
        self.conversao_csosn = {"102": "00", "500": "60", "900": "90"}
        self.meses = {"jan": "01", "feb": "02", "mar": "03", "apr": "04", "may": "05", "jun": "06",
                      "jul": "07", "aug": "08", "sep": "09", "oct": "10", "nov": "11", "dec": "12"}

    def _to_int_str(self, valor):
        if valor is None or (isinstance(valor, float) and valor != valor):
            return ""
        if isinstance(valor, float):
            return str(int(valor))
        return str(valor).strip()

    def _normalizar(self, valor):
        if valor is None or (isinstance(valor, float) and valor != valor):
            return ""
        if isinstance(valor, float):
            if valor == int(valor):
                return int(valor)
            return round(valor, 4)
        return str(valor).strip()

    def _fmt(self, valor, campo):
        if valor is None or (isinstance(valor, float) and valor != valor):
            return ""
        if isinstance(valor, float):
            if campo in ("procod", "clf_ncm", "procest", "proorig"):
                return str(int(valor))
            if valor == int(valor):
                return int(valor)
            return round(valor, 4)
        if campo == "procsosn":
            return self.conversao_csosn.get(str(valor), str(valor))
        return str(valor).strip()

    def _fmt_cli(self, v):
        if v is None or (isinstance(v, float) and v != v):
            return ""
        if isinstance(v, float):
            if v == int(v):
                return int(v)
            return round(v, 2)
        return str(v).strip()

    def _conv_data(self, s):
        if not s or s == "":
            return ""
        s = str(s).strip().lower()
        try:
            parts = s.replace("-", " ").split()
            if len(parts) == 3:
                d, m, a = parts
                if m in self.meses:
                    a_full = "20" + a if len(a) == 2 else a
                    return f"{d}/{self.meses[m]}/{a_full}"
        except Exception:
            pass
        return ""

    def _read_xls(self, path: Path):
        book = xlrd.open_workbook(str(path), encoding_override="latin1")
        sheet = book.sheet_by_index(0)
        headers = [sheet.cell_value(0, j) for j in range(sheet.ncols)]
        rows = []
        for i in range(1, sheet.nrows):
            row = {}
            for j, header in enumerate(headers):
                cell = sheet.cell(i, j)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    row[header] = None
                elif cell.ctype == xlrd.XL_CELL_TEXT:
                    row[header] = sheet.cell_value(i, j).strip()
                else:
                    row[header] = sheet.cell_value(i, j)
            rows.append(row)
        return rows

    def convert_products(self, acsn_path: Path, template_path: Path, output_name: str) -> Path:
        dados = self._read_xls(acsn_path)
        df = pd.DataFrame(dados)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = BASE / "output"
        output_dir.mkdir(exist_ok=True)
        saida = output_dir / f"{output_name.replace('.xlsx', '')}_{ts}.xlsx"
        shutil.copy(str(template_path), str(saida))

        wb = load_workbook(str(saida))
        ws = wb.active

        mapeamento = {
            2: "procod", 3: "prodes", 5: "prouni",
            6: "clf_ncm", 7: "procest", 8: "propdv", 9: "procsosn",
            12: "propdc", 14: "proref", 17: "propeso",
            24: "propis", 25: "prostipi", 26: "procodipi",
            28: "proorig", 29: "proipi", 35: "proqmi", 37: "proipe",
        }

        def _is_gtin(v):
            s = str(v).strip().lstrip("0")
            return s.isdigit() and len(s) in (8, 12, 13, 14)

        linha_inicio = 3
        for idx, row in df.iterrows():
            for col_num, campo in mapeamento.items():
                v = row.get(campo)
                ws.cell(row=linha_inicio + idx, column=col_num, value=self._fmt(v, campo))
            procod_val = self._fmt(row.get("procod"), "procod")
            if _is_gtin(procod_val):
                ws.cell(row=linha_inicio + idx, column=4, value=procod_val)

        wb.save(str(saida))
        return saida

    def analyze_clients(self, client_path: Path, supplier_path: Path = None):
        dados = []
        for r in self._read_xls(client_path):
            r["_tipo"] = "C"
            dados.append(r)
        if supplier_path:
            for r in self._read_xls(supplier_path):
                r["_tipo"] = "F"
                dados.append(r)

        sem_cep = []
        sem_cidade = []
        sem_estado = []
        amostras = {"sem_cep": [], "sem_cidade": [], "sem_estado": []}

        for idx, row in enumerate(dados):
            tipo = row.get("_tipo", "")
            cod = self._fmt_cli(row.get("fccod", ""))
            nome = self._fmt_cli(row.get("fcnom", ""))
            cep = self._fmt_cli(row.get("fccep", ""))
            cidade = self._fmt_cli(row.get("fccid", ""))
            estado = self._fmt_cli(row.get("fcest", ""))

            if not cep:
                sem_cep.append(idx)
                if len(amostras["sem_cep"]) < 5:
                    amostras["sem_cep"].append({"codigo": cod, "nome": nome, "tipo": tipo})

            if not cidade:
                sem_cidade.append(idx)
                if len(amostras["sem_cidade"]) < 5:
                    amostras["sem_cidade"].append({"codigo": cod, "nome": nome, "tipo": tipo})

            if not estado:
                sem_estado.append(idx)
                if len(amostras["sem_estado"]) < 5:
                    amostras["sem_estado"].append({"codigo": cod, "nome": nome, "tipo": tipo})

        return {
            "total": len(dados),
            "sem_cep": {"count": len(sem_cep), "amostras": amostras["sem_cep"]},
            "sem_cidade": {"count": len(sem_cidade), "amostras": amostras["sem_cidade"]},
            "sem_estado": {"count": len(sem_estado), "amostras": amostras["sem_estado"]},
        }

    def convert_clients(self, client_path: Path, template_path: Path, output_name: str, supplier_path: Path = None, fixes: dict = None) -> Path:
        dados = []
        for r in self._read_xls(client_path):
            r["_tipo"] = "C"
            dados.append(r)
        if supplier_path:
            for r in self._read_xls(supplier_path):
                r["_tipo"] = "F"
                dados.append(r)

        fixes = fixes or {}
        default_cep = (fixes.get("default_cep") or "").strip()
        default_city = (fixes.get("default_city") or "").strip()
        default_state = (fixes.get("default_state") or "").strip()
        exclude_missing = fixes.get("exclude_missing_city", False)

        if default_cep:
            for row in dados:
                if not self._fmt_cli(row.get("fccep", "")):
                    row["fccep"] = default_cep
        if default_city:
            for row in dados:
                if not self._fmt_cli(row.get("fccid", "")):
                    row["fccid"] = default_city
        if default_state:
            for row in dados:
                if not self._fmt_cli(row.get("fcest", "")):
                    row["fcest"] = default_state

        if exclude_missing:
            dados = [r for r in dados if self._fmt_cli(r.get("fccid", "")) and self._fmt_cli(r.get("fcest", ""))]

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = BASE / "output"
        output_dir.mkdir(exist_ok=True)
        saida = output_dir / f"{output_name.replace('.xlsx', '')}_{ts}.xlsx"
        shutil.copy(str(template_path), str(saida))

        wb = load_workbook(str(saida))
        ws = wb.active

        linha_inicio = 4
        for idx, row in enumerate(dados):
            cep = self._fmt_cli(row.get("fccep", ""))
            cidade = self._fmt_cli(row.get("fccid", ""))
            estado = self._fmt_cli(row.get("fcest", ""))
            ws.cell(row=linha_inicio + idx, column=2, value=self._fmt_cli(row.get("fccod")))
            ws.cell(row=linha_inicio + idx, column=3, value=self._fmt_cli(row.get("fcnom")).upper())
            ws.cell(row=linha_inicio + idx, column=4, value=self._fmt_cli(row.get("fcnom")).upper())
            ws.cell(row=linha_inicio + idx, column=5, value=self._fmt_cli(row.get("fccpg")))
            ws.cell(row=linha_inicio + idx, column=6, value=row.get("_tipo", ""))
            ws.cell(row=linha_inicio + idx, column=7, value=self._fmt_cli(row.get("fcrg")))
            ws.cell(row=linha_inicio + idx, column=9, value=self._fmt_cli(row.get("fcend")))
            ws.cell(row=linha_inicio + idx, column=10, value="SN")
            ws.cell(row=linha_inicio + idx, column=11, value=self._fmt_cli(row.get("fccom")))
            ws.cell(row=linha_inicio + idx, column=12, value=cep)
            ws.cell(row=linha_inicio + idx, column=13, value=self._fmt_cli(row.get("fcbai")))
            ws.cell(row=linha_inicio + idx, column=15, value=cidade.upper())
            ws.cell(row=linha_inicio + idx, column=14, value=self.ibge.get(cidade, estado))
            ws.cell(row=linha_inicio + idx, column=16, value=estado)
            data_cad = self._conv_data(row.get("fccad", ""))
            if data_cad:
                ws.cell(row=linha_inicio + idx, column=17, value=data_cad)
                ws.cell(row=linha_inicio + idx, column=18, value=data_cad)
            else:
                ws.cell(row=linha_inicio + idx, column=17, value="01/01/1990")
                ws.cell(row=linha_inicio + idx, column=18, value="01/01/1990")
            ws.cell(row=linha_inicio + idx, column=19, value=self._fmt_cli(row.get("fcemail")))
            ws.cell(row=linha_inicio + idx, column=20, value=self._fmt_cli(row.get("fcfon")))
            ws.cell(row=linha_inicio + idx, column=23, value=self._fmt_cli(row.get("fclimcr")))
            ws.cell(row=linha_inicio + idx, column=25, value=data_cad if data_cad else "01/01/1990")

        wb.save(str(saida))
        return saida
